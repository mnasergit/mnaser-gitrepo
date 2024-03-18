import openpyxl
from openpyxl import Workbook
from ipaddress import IPv4Interface

from nornir import InitNornir
from nornir_napalm.plugins.tasks import napalm_get
from nornir_utils.plugins.functions import print_result
from nornir_netmiko.tasks import netmiko_send_command
import json
import pprint

num_group = 5

# Open Workbook and Sheet 
wb = openpyxl.load_workbook("Device-Info.xlsx")
sheet = wb['IPv6_Basic']

max_row = num_group
max_column = sheet.max_column

hosts_config = ""

for i in range(2, num_group + 2):
    hostname = sheet.cell(row=i, column=1).value
    IPv4_FA60_Temp = sheet.cell(row=i,column=3).value
    IPv4_FA60_IP = IPv4Interface(IPv4_FA60_Temp)
    IPv4_FA60_IP = str(IPv4_FA60_IP.ip)

    hosts_config += f"{hostname}:\n"
    hosts_config += f"    hostname: {IPv4_FA60_IP}\n"
    hosts_config += f"    platform: 'ios'\n"
    hosts_config += f"    username: apnic\n"
    hosts_config += f"    password: training\n"
    hosts_config += f"\n"

# Write startup-config in plain-text for individual router 
with open(f"hosts.yaml", 'w') as write_cfg:
    write_cfg.seek(0)
    write_cfg.write(hosts_config)
    
nr = InitNornir(
    config_file="hosts.yaml", dry_run=True
)

#print(nr.inventory.hosts)

results = nr.run(
    task=napalm_get, getters=["facts", "interfaces", "interfaces_ip"]
    #task=napalm_get, getters=["config"]
)


# Convert each individual result to a dictionary
results_dict = {host: result[0].result for host, result in results.items()}

### If it is required to see the file ###
# Convert OrderedDict to JSON string
result_json = json.dumps(results_dict)

# # Convert JSON string to pprint format
result_json = json.loads(result_json)
result_json = pprint.pformat(result_json)

with open(f"results_dict.json", "w") as result_file:
     result_file.write(result_json)


for router_name, router_data in results_dict.items():
    print(router_name)
    print("=" * len(router_name))
    hostname = router_data['facts']['hostname']
    mgmt_ip_data = router_data['interfaces_ip']['FastEthernet6/0']['ipv4']
    print(f"MGMT IP: {mgmt_ip_data}")
