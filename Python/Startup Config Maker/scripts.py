import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import ipaddress
from ipaddress import IPv6Interface
from ipaddress import IPv4Interface

# Open Workbook and Sheet 
wb = openpyxl.load_workbook("Device-Info.xlsx")
sheet = wb['Sheet1']

# Define Function
def update_router_config(config_text):
    new_config = ""
    lines = config_text.split('\n')
    replace_flag = False

    for line in lines:
        if line.strip().startswith("hostname Router"):
            replace_flag = True
            new_config += "hostname " + hostname + "\n"
            new_config += "!\n"
        elif line.strip().startswith("interface FastEthernet6/0"):
            replace_flag = True
            new_config += "interface FastEthernet6/0\n"
            new_config += " description **MGMT - Do Not Delete**\n"
            new_config += " ipv6 address " + str(IPv6_FA60_IP.ip) + "/" + str(IPv6_FA60_IP._prefixlen) + "\n"
            new_config += " ip address " + str(IPv4_FA60_IP.ip) + " " + str(IPv4_FA60_IP.netmask) + "\n"
            new_config += " no shutdown\n"
            new_config += "!\n"
        elif line.strip().startswith("line con 0"):
            replace_flag = True
            new_config += "line con 0\n"
            new_config += " exec-timeout 0 0\n"
            new_config += " logging synchronous\n"
            new_config += "!\n"
        elif line.strip().startswith("line vty 0 4"):
            replace_flag = True
            new_config += "line vty 0 4\n"
            new_config += " exec-timeout 0 0\n"
            new_config += " logging synchronous\n"
            new_config += " login local\n"
            new_config += " transport input all\n"
            new_config += "!\n"
        elif line.strip().startswith("end"):
            replace_flag = True
            new_config += "!\n"
        elif replace_flag and line.strip().startswith("!"):
            replace_flag = False
        elif not replace_flag:
            new_config += line + "\n"
    
    return new_config

# Iterating Function
try:
    max_row = sheet.max_row
    max_column = sheet.max_column

    for i in range(2, max_row + 1):
        hostname = sheet.cell(row=i, column=1).value
        IPv6_FA60_Temp = sheet.cell(row=i, column=3).value
        IPv6_FA60_IP = IPv6Interface(IPv6_FA60_Temp)
        IPv4_FA60_Temp = sheet.cell(row=i,column=4).value
        IPv4_FA60_IP = IPv4Interface(IPv4_FA60_Temp)

        # Read default configuration of router
        with open("default_config.txt", 'r') as default_cfg:
            router_config = default_cfg.read()
            # Update the router configuration
            updated_config = update_router_config(router_config)
        
        # Read common configuration of router
        with open("common_config.txt", 'r') as common_cfg:
            # Update the common configuration
            common_config = common_cfg.read()

        # Write configuration in individual file for each router 
        with open(f"{hostname}.txt", 'w') as write_cfg:
            write_cfg.seek(0)
            write_cfg.write(updated_config)
            write_cfg.write(common_config)

    wb.close()
    default_cfg.close()
    common_cfg.close()
    write_cfg.close()

# Handling Error            
except ValueError:
    print ()
    print((f"Invalid input deceted in file: *** {wb} *** --- "), ipaddress)
    print ()
    wb.close()