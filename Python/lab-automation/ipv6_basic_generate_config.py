import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import ipaddress
from ipaddress import IPv6Interface
#from ipaddress import IPv6Address
#from ipaddress import IPv6Network
from ipaddress import IPv4Interface
#from ipaddress import IPv4Address
#from ipaddress import IPv4Network
import base64
import time
import json
import sys
import app
import re

def process_variable(input_value):
    #print(f"Processing input value: {input_value}")
    # Your processing logic here
    #num_group = f"Processed variable: {input_value}"
    #num_group = int(input_value)
    #print(f"Result: {num_group}")
    #return num_group

    match = re.search(r'\b\d+\b', input_value)
    if match:
        num_group = match.group()
        return num_group
    else:
        return "0"  # Return a default value if no numeric part is found

if __name__ == "__main__":
    try:
        # Get the input value from the command line argument
        input_value = sys.argv[1]
        num_group = process_variable(input_value)
        num_group = int(num_group)
        # Rest of your existing code...

        ### User input ###
        #num_group = input("Enter nmber of groups for IPv6 Basic Lab (1-20): ")

        # try:
        #     num_group = int(num_group)
        #     #num_group = int(process_variable(input_value))
        # # Handling Error
        # except:
        #     print ("Invalid input detected for < Group Number >")

        try:
            if num_group > 0 and  num_group <= 20:
                # Open Workbook and Sheet 
                wb = openpyxl.load_workbook("Device-Info.xlsx")
                sheet = wb['IPv6_Basic']

                # Create config from xlsx file
                def update_router_config(config_text):
                    new_config = ""
                    lines = config_text.split('\n')
                    replace_flag = False

                    for line in lines:
                        if line.strip().startswith("hostname Router"):
                            replace_flag = True
                            new_config += "hostname " + hostname + "\n"
                            new_config += "!\n"
                        elif line.strip().startswith("interface FastEthernet6/0"):
                            replace_flag = True
                            new_config += "interface FastEthernet6/0\n"
                            new_config += " description **MGMT - Do Not Delete**\n"
                            new_config += " ip address " + str(IPv4_FA60_IP.ip) + " " + str(IPv4_FA60_IP.netmask) + "\n"
                            new_config += " duplex full\n"
                            new_config += " ipv6 address " + str(IPv6_FA60_IP.ip) + "/" + str(IPv6_FA60_IP._prefixlen) + "\n"
                            new_config += "!\n"
                        elif line.strip().startswith("line con 0"):
                            replace_flag = True
                            new_config += "line con 0\n"
                            new_config += " exec-timeout 0 0\n"
                            new_config += " logging synchronous\n"
                            new_config += "!\n"
                        elif line.strip().startswith("line vty 0 4"):
                            replace_flag = True
                            new_config += "line vty 0 4\n"
                            new_config += " exec-timeout 0 0\n"
                            new_config += " logging synchronous\n"
                            new_config += " login local\n"
                            new_config += " transport input all\n"
                            new_config += "!\n"
                        elif line.strip().startswith("end"):
                            replace_flag = True
                            new_config += "!\n"
                        elif replace_flag and line.strip().startswith("!"):
                            replace_flag = False
                        elif not replace_flag:
                            new_config += line + "\n"
                    return new_config
                    
                # Create task-1 config xlsx file
                def update_task1_config(config_task1_text):
                    new_task1_config = ""
                    lines = config_task1_text.split('\n')
                    replace_flag = False

                    for line in lines:
                        if line.strip().startswith("interface FastEthernet0/0"):
                            replace_flag = True
                            new_task1_config += "interface FastEthernet0/0\n"
                            new_task1_config += " description **Connected to Host1**\n"
                            new_task1_config += " no ip address\n"
                            new_task1_config += " duplex full\n"
                            new_task1_config += " ipv6 address 2001:db8:abc::1/64\n"
                            new_task1_config += "!\n"
                        elif replace_flag and line.strip().startswith("!"):
                            replace_flag = False
                        elif not replace_flag:
                            new_task1_config += line + "\n"
                    return new_task1_config    
                    
                # Interate and create config from default config and common config file
                try:
                    max_row = num_group
                    max_column = sheet.max_column

                    for i in range(2, num_group + 2):
                        hostname = sheet.cell(row=i, column=1).value
                        IPv6_FA60_Temp = sheet.cell(row=i, column=2).value
                        IPv6_FA60_IP = IPv6Interface(IPv6_FA60_Temp)
                        IPv4_FA60_Temp = sheet.cell(row=i,column=3).value
                        IPv4_FA60_IP = IPv4Interface(IPv4_FA60_Temp)

                        json_data = {"id": "1", "data": ""}

                        # Read default config file of router
                        with open("default_config.txt", 'r') as default_cfg:
                            router_config = default_cfg.read()
                            # Update the router configuration
                            updated_config = update_router_config(router_config)
                        
                        # Read common config file of router
                        with open("common_config.txt", 'r') as common_cfg:
                            # Update the common configuration
                            common_config = common_cfg.read()

                        # Write startup-config in plain-text for individual router 
                        with open(f"ipv6_basic_config_files/{hostname}-startup-config.txt", 'w') as write_cfg:
                            write_cfg.seek(0)
                            write_cfg.write(updated_config)
                            write_cfg.write(common_config)
                            #print(f"{hostname} : Startup-config file (plain) created.")

                        # # Converting startup-config from plain-text to hash
                        # with open(f"ipv6_basic_out_files/{hostname}-startup-config.txt", "rb") as text_file:
                        #     #print (f"Reading {hostname}-startup-config.txt")
                        #     text_content = text_file.read()
                            
                        #     # Encode contents using base64
                        #     encoded_contents = base64.b64encode(text_content)

                        # # Write encoded contents to a new file
                        # with open(f"ipv6_basic_out_files/{hostname}-startup-config-hashed.txt", 'wb') as hash_file:
                        #     #print (f"Writing hash for {hostname}-startup-config.txt")
                        #     hash_file.write(encoded_contents)
                        #     #print(f"{hostname} : Startup-config file (hash) created.")
                            
                        # text_file.close()
                        # hash_file.close()

                        # # Converting startup-config from plain-text to json
                        # with open(f"ipv6_basic_out_files/{hostname}-startup-config.txt", "r") as text_file:
                        #     json_data["data"] = text_file.read()
                        
                        #     # Convert the Python dictionary to a JSON string
                        #     json_string = json.dumps(json_data, indent=2)

                        # with open(f"ipv6_basic_out_files/{hostname}-startup-config-json.txt", 'w') as json_file:
                        #     json_file.write(json_string)
                        #     #print(f"{hostname} : Startup-config file (json) created.")
                        #     #print("")
                        
                        # text_file.close()
                        # json_file.close()

                        default_cfg.close()
                        common_cfg.close()
                        write_cfg.close()
                        
                        ### Create Task-1 Config for all router ###
                        with open(f"ipv6_basic_config_files/{hostname}-startup-config.txt", 'r') as start_cfg:
                            router_config = start_cfg.read()
                            # Update the router configuration
                            updated_config = update_task1_config(router_config)

                        # Write task1-config in plain-text for individual router 
                        with open(f"ipv6_basic_config_files/{hostname}-task1-config.txt", 'w') as write_cfg:
                            write_cfg.seek(0)
                            write_cfg.write(updated_config)
                            #print(f"{hostname} : Startup-config file (plain) created.")

                        start_cfg.close()
                        write_cfg.close()
                    
                # Handling Error            
                except ValueError:
                    print(f"Invalid input detected in XLSX file: < {wb} >")
                    wb.close()

                wb.close()
                print(f"All config files have been genrated in directory: < /ipv6_basic_config_files >.")
                print("")
                
            else:
                print("Invalid input detected for < Group Number >")
        
        # Handling Error
        except ValueError:
            print("Error: Invalid input. Please provide a valid integer value for the number of groups.")
            sys.exit(1)

    #except IndexError:
     #   print("Error: Please provide the input value.")
    except IndexError:
        print("Error: Please provide the input value.")
