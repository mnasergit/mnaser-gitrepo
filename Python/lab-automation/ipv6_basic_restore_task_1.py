import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from ipaddress import IPv4Interface
import requests
import json
from pprint import pprint
import time
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor
from netmiko import ConnectHandler
import sqlite3
from lab_variable import EVE_1, EVE_2, EVE_USER, EVE_PASSWORD, ROUTER_USER, ROUTER_PASSWORD, num_group_eve_1


# Connect to the same SQLite database
conn = sqlite3.connect('lab_data.db')
cursor = conn.cursor()

# Retrieve the value from the table
cursor.execute("SELECT value FROM table_basic_ipv6")
input_value = cursor.fetchone()[0]
lab_name = "IPv6-Basic-Connectivity-Lab"
lab_name_check = f"/{lab_name}.unl"
num_group = int(input_value)
num_group_eve_2 = num_group - num_group_eve_1
num_ios = num_group
num_slax = num_group
num_ios_eve_1 = num_group_eve_1
num_slax_eve_1 = num_group_eve_1
num_ios_eve_2 = num_group_eve_2
num_slax_eve_2 = num_group_eve_2
headers = {"Accept":"application/json"}

def eve_ng_login():
    cookies1 = None
    cookies2 = None
    
    ### Login into EVE1 ###
    try:
        login_url = f"http://{EVE_1}/api/auth/login"
        cred = '{{"username":"{}","password":"{}","html5":"-1"}}'.format(EVE_USER, EVE_PASSWORD)

        login_api = requests.post(url=login_url, data=cred, headers=headers)
        cookies1 = login_api.cookies
    
    except:
        print("Login into EVE-NG-1 is not successful.")

    ### Login into EVE2 ###
    if num_group_eve_2 > 0:
        try:
            login_url = f"http://{EVE_2}/api/auth/login"
            cred = '{{"username":"{}","password":"{}","html5":"-1"}}'.format(EVE_USER, EVE_PASSWORD)

            login_api = requests.post(url=login_url, data=cred, headers=headers)
            cookies2 = login_api.cookies
        
        except:
            print("Login into EVE-NG-2 is not successful.")

    return cookies1, cookies2

def task1_restore1():
    # Open Workbook and Sheet
    wb = openpyxl.load_workbook("Device-Info.xlsx")
    sheet = wb['IPv6_Basic']

    # Iterating Function
    max_row = sheet.max_row
    max_column = sheet.max_column
    HostnameList = []
    IPList = []

    for i in range(2, int(num_ios) + 2):
        hostname = sheet.cell(row=i, column=1).value
        IPv4_FA60_Temp = sheet.cell(row=i,column=3).value
        IPv4_FA60_IP = IPv4Interface(IPv4_FA60_Temp)
        IP = str(IPv4_FA60_IP.ip)
        HostnameList.append(hostname)
        IPList.append(IP)

    # Create an empty report file
    with open("ipv6_basic_report/Lab-Report-Task1-Restore.txt", "w"):
        pass

    return HostnameList, IPList, hostname, IP


def task1_restore2(hostname, IP):
    result = f"\n"
    try:
        cisco_devices = {
            'device_type': 'cisco_ios',
            'host': IP,
            'username': ROUTER_USER,
            'password': ROUTER_PASSWORD
        }

        with ConnectHandler(**cisco_devices) as ssh:
            #print(f"Executing script on {HOST}...")
            command1 = (f"configure replace tftp://10.99.99.11/ipv6_basic_config_files/{hostname}-task1-config.txt")
            command2 = "Y"
            command3 = "write memory"
            output1 = ssh.send_command_timing(command1, strip_prompt=False, strip_command=False)
            time.sleep(3)
            #print (output1)
            output2 = ssh.send_command_timing(command2, strip_prompt=False, strip_command=False)
            time.sleep(6)
            #print (output2)
            output3 = ssh.send_command_timing(command3, strip_prompt=False, strip_command=False)
            time.sleep(3)
            #print (output3)
            result += f"{hostname}: IPv6 address configuration is restored successfully."
            result += f"\n" 

    except Exception as e:
        result += f"Couldn't connect to {hostname}, check manually!"
        result += f"\n"

    return result

def push_startup_config(cookies1, cookies2):
    ### Push startup-configs in EVE1 ###
    j = 1
    for i in range (1, num_ios_eve_1 + 1):    
        # Read the contents of the text file
        json_data = {"id": "1", "data": ""}

        with open(f"ipv6_basic_config_files/Group{i}-Router1-task1-config.txt", "r") as file:
            #print (f"Reading {HOST}-startup-config.txt")
            json_data["data"] = file.read()

            # Convert the Python dictionary to a JSON string
            json_string = json.dumps(json_data, indent=2)

            # Save the JSON string to a new file
            #with open('output.json', 'w') as json_file:
            #    json_file.write(json_string)

            ### Push Config ###
            push_config_data = json_string

            #print ("Lab data: " + lab_data)
            #lab_data = json.dumps(lab_data)
                    
            push_config_url = f"http://{EVE_1}/api/labs/{lab_name_check}/configs/{j}"
            push_config_api = requests.put(url=push_config_url,data=push_config_data,cookies=cookies1,headers=headers)
            
            ### Set Config Bit in EVE1 ###
            config_bit_data = {"id":1,"count":1,"postfix":0,"config":1}
            config_bit_data = json.dumps(config_bit_data)
                    
            config_bit_url = f"http://{EVE_1}/api/labs/{lab_name_check}/nodes/{j}"
            config_bit_api = requests.put(url=config_bit_url,data=config_bit_data,cookies=cookies1,headers=headers)
    
            j = j + 1

    ### Push startup-configs in EVE2 ###
    j = 1
    for i in range (num_ios_eve_1 + 1, num_ios + 1):    
        # Read the contents of the text file
        json_data = {"id": "1", "data": ""}

        with open(f"ipv6_basic_config_files/Group{i}-Router1-startup-config.txt", "r") as file:
            #print (f"Reading {HOST}-startup-config.txt")
            json_data["data"] = file.read()

            # Convert the Python dictionary to a JSON string
            json_string = json.dumps(json_data, indent=2)

            # Save the JSON string to a new file
            #with open('output.json', 'w') as json_file:
            #    json_file.write(json_string)

            ### Push Config ###
            push_config_data = json_string

            #print ("Lab data: " + lab_data)
            #lab_data = json.dumps(lab_data)
                    
            push_config_url = f"http://{EVE_2}/api/labs/{lab_name_check}/configs/{j}"
            push_config_api = requests.put(url=push_config_url,data=push_config_data,cookies=cookies2,headers=headers)
            ### Set Config Bit in EVE2 ###
            config_bit_data = {"id":1,"count":1,"postfix":0,"config":1}
            config_bit_data = json.dumps(config_bit_data)
                    
            config_bit_url = f"http://{EVE_2}/api/labs/{lab_name_check}/nodes/{j}"
            config_bit_api = requests.put(url=config_bit_url,data=config_bit_data,cookies=cookies2,headers=headers)
            j = j + 1


def main():
    cookies1, cookies2 = eve_ng_login()
    HostnameList, IPList, hostname, IP = task1_restore1()

    with ThreadPoolExecutor(max_workers=len(HostnameList)) as executor:
        results = list(executor.map(task1_restore2, HostnameList, IPList))
    
    # Write the results to the report file in the correct order
    with open("ipv6_basic_report/Lab-Report-Task1-Restore.txt", "a") as savereport:
        for result in results:
            savereport.write(result)

    # Show the report content
    with open("ipv6_basic_report/Lab-Report-Task1-Restore.txt", "r") as showreport:
        print(showreport.read())

    # Call the export_node function
    push_startup_config(cookies1, cookies2)

if __name__ == '__main__':
    main()
