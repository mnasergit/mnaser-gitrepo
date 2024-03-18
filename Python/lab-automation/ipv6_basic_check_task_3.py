from netmiko import ConnectHandler
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from ipaddress import IPv4Interface
import requests
from pprint import pprint
import time
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor
from lab_variable import EVE_1, EVE_USER, EVE_PASSWORD, ROUTER_USER, ROUTER_PASSWORD

lab_name = "IPv6-Basic-Connectivity-Lab"
lab_name_check = f"/{lab_name}.unl"

### Login ###
login_url = f"http://{EVE_1}/api/auth/login"
cred = '{{"username":"{}","password":"{}","html5":"-1"}}'.format(EVE_USER, EVE_PASSWORD)
headers = {"Accept":"application/json"}

login_api = requests.post(url=login_url, data=cred)
cookies = login_api.cookies

# Check Status of Node #
node_status_url = f"http://{EVE_1}/api/labs/IPv6-Basic-Connectivity-Lab.unl/nodes"
node_status_api = requests.get(url=node_status_url,cookies=cookies,headers=headers)

node_status_api_response = node_status_api.json()
#pprint (node_status_api_response)

node_status_dict = node_status_api_response['data']
num_nodes = len(node_status_dict)

# Open Workbook and Sheet 
wb = openpyxl.load_workbook("Device-Info.xlsx")
sheet = wb['IPv6_Basic']

# Iterating Function
max_row = sheet.max_row
max_column = sheet.max_column
HostnameList = []
IPList = []

for i in range(2, int(num_nodes/2)+2):
    hostname = sheet.cell(row=i, column=1).value
    IPv4_FA60_Temp = sheet.cell(row=i,column=3).value
    IPv4_FA60_IP = IPv4Interface(IPv4_FA60_Temp)
    IP = str(IPv4_FA60_IP.ip)   
    HostnameList.append(hostname)
    IPList.append(IP)

# Create an empty report file
with open("ipv6_basic_report/Lab-Report-Task3-Check.txt", "w"):
    pass

def task3_check(hostname, IP):
    result = ""
    try:
        cisco_devices = {
            'device_type': 'cisco_ios',
            'host': IP,
            'username': ROUTER_USER,
            'password': ROUTER_PASSWORD
        }
    
        with ConnectHandler(**cisco_devices) as ssh:
            command1 = f"ping 2001:db8:abc::2 repeat 3"
            output1 = ssh.send_command(command1, strip_prompt=False, strip_command=False)
            time.sleep(8)

            # Open the file in write mode and write the output to it
            with open(f"ipv6_basic_report/{hostname}-Verify-Connectivity.txt", "w") as saveoutput:
                saveoutput.write(output1)
                saveoutput.write("\n")

            with open(f"ipv6_basic_report/{hostname}-Verify-Connectivity.txt", "r") as readoutput:
                text = readoutput.read()
            
            RateStart = "Success rate is"
            RateEnd = "percent"

            if RateStart in text and RateEnd in text:
                RateStartIndex = text.index(RateStart)
                RateStartIndexLen = (len(RateStart))
                RateEndIndex = text.index(RateEnd)
                SuccessRate = text[RateStartIndex+RateStartIndexLen:RateEndIndex]
                
                SuccessRate = SuccessRate.strip(" ")
                SuccessRate = int(SuccessRate)
                
                result += f"\n"
                
                if SuccessRate > 0:
                    result += f"{hostname}: can reach host.\n"
                elif SuccessRate == 0: 
                    result += f"{hostname}: can't reach host...(!)\n"
                else:
                    result += f"{hostname}: couldn't read data"

    except Exception as e:
        result += f"\n"
        result += f"Couldn't connect to {hostname}, check manually...(!)"

    return result

def main():
    with ThreadPoolExecutor(max_workers=len(HostnameList)) as executor:
        results = list(executor.map(task3_check, HostnameList, IPList))

    # Write the results to the report file in the correct order
    with open("ipv6_basic_report/Lab-Report-Task3-Check.txt", "a") as savereport:
        for result in results:
            savereport.write(result)

    # Show the report content
    with open("ipv6_basic_report/Lab-Report-Task3-Check.txt", "r") as showreport:
        print(showreport.read())

if __name__ == '__main__':
    main()

