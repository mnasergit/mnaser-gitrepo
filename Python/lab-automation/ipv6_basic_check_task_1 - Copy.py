import openpyxl
import ipaddress
import requests
import time
from concurrent.futures import ThreadPoolExecutor
from netmiko import ConnectHandler
from lab_variable import EVE_1, EVE_USER, EVE_PASSWORD, ROUTER_USER, ROUTER_PASSWORD

lab_name = "IPv6-Basic-Connectivity-Lab"

# Login

login_url = f"http://{EVE_1}/api/auth/login"
cred = '{{"username":"{}","password":"{}","html5":"-1"}}'.format(EVE_USER, EVE_PASSWORD)
headers = {"Accept": "application/json"}

login_api = requests.post(url=login_url, data=cred)
cookies = login_api.cookies

#if login_api:
#    print("Login successful")
#else:
#    print("Login failed")
#    sys.exit(1)

# Check Status of Node
node_status_url = f"http://{EVE_1}/api/labs/{lab_name}.unl/nodes"
node_status_api = requests.get(url=node_status_url, cookies=cookies, headers=headers)

#if node_status_api:
#    print("Node status retrieved successfully")
#else:
#    print("Failed to retrieve node status")
#    sys.exit(1)

node_status_dict = node_status_api.json()['data']
num_nodes = len(node_status_dict)

# Open Workbook and Sheet
wb = openpyxl.load_workbook("Device-Info.xlsx")
sheet = wb['IPv6_Basic']

# Iterating Function
max_row = sheet.max_row
max_column = sheet.max_column
HostnameList = []
IPList = []

for i in range(2, int(num_nodes / 2) + 2):
    hostname = sheet.cell(row=i, column=1).value
    IPv4_FA60_Temp = sheet.cell(row=i, column=3).value
    IPv4_FA60_IP = ipaddress.IPv4Interface(IPv4_FA60_Temp)
    IP = str(IPv4_FA60_IP.ip)
    HostnameList.append(hostname)
    IPList.append(IP)

# Create an empty report file
with open("ipv6_basic_report/Lab-Report-Task1-Check.txt", "w"):
    pass

def task1_check(hostname, IP):
    result = ""
    try:
        cisco_devices = {
            'device_type': 'cisco_ios',
            'host': IP,
            'username': ROUTER_USER,
            'password': ROUTER_PASSWORD
        }

        with ConnectHandler(**cisco_devices) as ssh:
            command1 = f"show ipv6 interface brief fastEthernet0/0"
            output1 = ssh.send_command(command1, strip_prompt=False, strip_command=False)
            time.sleep(1)
            command2 = f"show ipv6 interface fastEthernet0/0"
            output2 = ssh.send_command(command2, strip_prompt=False, strip_command=False)
            time.sleep(1)

            with open(f"ipv6_basic_report/{hostname}-IPv6-Config-Check.txt", "w") as saveoutput:
                saveoutput.write(output1)
                saveoutput.write("\n")
                saveoutput.write(output2)
                saveoutput.write("\n")

            with open(f"ipv6_basic_report/{hostname}-IPv6-Config-Check.txt", "r") as readoutput:
                text = readoutput.read()

            IntStatusStart = "["
            IntStatusEnd = "]"

            if IntStatusStart in text and IntStatusEnd in text:
                IntStatusStartIndex = text.index(IntStatusStart)
                IntStatusEndIndex = text.index(IntStatusEnd)
                IntStatus = text[IntStatusStartIndex + 1:IntStatusEndIndex]
            else:
                print(f"Couldn't read Interface Status of {hostname}")

            GUAStart = "Global unicast address"
            GUAEnd = "Joined group address"

            if GUAStart in text and GUAEnd in text:
                GUAStartIndex = text.index(GUAStart)
                GUAStartIndexLen = len(GUAStart)
                GUAEndIndex = text.index(GUAEnd)
                GUAAddress = text[GUAStartIndex + GUAStartIndexLen:GUAEndIndex]
                CIDRIndex = GUAAddress.index("/")
                CIDRvalue = GUAAddress[CIDRIndex + 1:CIDRIndex + 4]
                PrefixLenValue = int(CIDRvalue)

                GUA = "2001:DB8:ABC::1"
                GUACount = text.count("subnet is")

                result += f"\n"

                if f"{GUA}," in GUAAddress and GUACount == 1 and PrefixLenValue == 64:
                    result += f"{hostname} IPv6 address and prefix length are correct."
                elif f"{GUA}," in GUAAddress and GUACount == 1 and PrefixLenValue != 64:
                    result += f"{hostname} IPv6 address is correct but prefix length is not correct...(!)"
                elif f"{GUA}," in GUAAddress and GUACount > 1:
                    result += f"{hostname} has multiple IPv6 configured...(!)"
                else:
                    result += f"{hostname} IPv6 address is not correct...(!)"
                
                result += f"\n"

                if IntStatus == "up/up":
                    result += f"{hostname} interface Fa0/0 status is UP."
                else:
                    result += f"{hostname} interface Fa0/0 status is DOWN...(!)"
                        
                result += f"\n"

            else:
                result += f"\n{hostname} IPv6 address is not configured...(!)"
                result += f"\n"

    except Exception as e:
        result += f"\n{hostname} couldn't connect...(!)"
        result += f"\n"

    return result

def main():
    with ThreadPoolExecutor(max_workers=len(HostnameList)) as executor:
        results = list(executor.map(task1_check, HostnameList, IPList))

    # Write the results to the report file in the correct order
    with open("ipv6_basic_report/Lab-Report-Task1-Check.txt", "a") as savereport:
        for result in results:
            savereport.write(result)

    # Show the report content
    with open("ipv6_basic_report/Lab-Report-Task1-Check.txt", "r") as showreport:
        print(showreport.read())

if __name__ == '__main__':
    main()
