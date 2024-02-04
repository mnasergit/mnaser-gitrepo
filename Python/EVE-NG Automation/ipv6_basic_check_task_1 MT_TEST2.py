import openpyxl
from openpyxl.styles import PatternFill
import ipaddress
from ipaddress import IPv4Interface
import requests
import json
import time
import sys
from netmiko import ConnectHandler
import multiprocessing

EVE_1 = "192.168.20.12"
ROUTER_USER = "apnic"
ROUTER_PASSWORD = "training"

# User Input
lab_name = "IPv6-Basic-Connectivity-Lab"

# Login
login_url = f"http://{EVE_1}/api/auth/login"
cred = '{"username":"admin","password":"eve","html5":"-1"}'
headers = {"Accept": "application/json"}

login_api = requests.post(url=login_url, data=cred)
cookies = login_api.cookies

if login_api:
    print("Login successful")
else:
    print("Login failed")
    sys.exit(1)

# Check Status of Node
node_status_url = f"http://{EVE_1}/api/labs/{lab_name}.unl/nodes"
node_status_api = requests.get(url=node_status_url, cookies=cookies, headers=headers)

if node_status_api.ok:
    print("Node status retrieved successfully")
else:
    print("Failed to retrieve node status")
    sys.exit(1)

node_status_dict = node_status_api.json()['data']
num_nodes = len(node_status_dict)

# Open Workbook and Sheet
wb = openpyxl.load_workbook("Device-Info.xlsx")
sheet = wb['IPv6_Basic']

# Iterating Function
max_row = sheet.max_row
max_column = sheet.max_column
HostnameList = []
IPList = []

for i in range(2, int(num_nodes / 2) + 2):
    hostname = sheet.cell(row=i, column=1).value
    IPv4_FA60_Temp = sheet.cell(row=i, column=3).value
    IPv4_FA60_IP = IPv4Interface(IPv4_FA60_Temp)
    IP = str(IPv4_FA60_IP.ip)
    HostnameList.append(hostname)
    IPList.append(IP)

savereport = open("Lab-Report-Task1.txt", "w")


def task1_check(hostname, IP):
    try:
        cisco_devices = {
            'device_type': 'cisco_ios',
            'host': IP,
            'username': ROUTER_USER,
            'password': ROUTER_PASSWORD
        }

        with ConnectHandler(**cisco_devices) as ssh:
            command1 = f"show ipv6 interface brief fastEthernet0/0"
            output1 = ssh.send_command(command1, strip_prompt=False, strip_command=False)
            time.sleep(1)
            command2 = f"show ipv6 interface fastEthernet0/0"
            output2 = ssh.send_command(command2, strip_prompt=False, strip_command=False)
            time.sleep(1)

            with open(f"{hostname}-task1-compare.txt", "w") as saveoutput:
                saveoutput.write(output1)
                saveoutput.write("\n")
                saveoutput.write(output2)
                saveoutput.write("\n")

            with open(f"{hostname}-task1-compare.txt", "r") as readoutput:
                text = readoutput.read()

            IntStatusStart = "["
            IntStatusEnd = "]"

            if IntStatusStart in text and IntStatusEnd in text:
                IntStatusStartIndex = text.index(IntStatusStart)
                IntStatusEndIndex = text.index(IntStatusEnd)
                IntStatus = text[IntStatusStartIndex + 1:IntStatusEndIndex]
            else:
                print(f"Couldn't read Interface Status of {hostname}")

            GUAStart = "Global unicast address"
            GUAEnd = "Joined group address"

            if GUAStart in text and GUAEnd in text:
                GUAStartIndex = text.index(GUAStart)
                GUAStartIndexLen = len(GUAStart)
                GUAEndIndex = text.index(GUAEnd)
                GUAAddress = text[GUAStartIndex + GUAStartIndexLen:GUAEndIndex]
                CIDRIndex = GUAAddress.index("/")
                CIDRvalue = GUAAddress[CIDRIndex + 1:CIDRIndex + 4]
                PrefixLenValue = int(CIDRvalue)

                GUA = "2001:DB8:ABC::1"
                GUACount = text.count("subnet is")

                with open("Lab-Report-Task1.txt", "a") as savereport:
                    savereport.write("\n")

                    if GUA in GUAAddress and GUACount == 1 and PrefixLenValue == 64:
                        savereport.write(f"{hostname} IPv6 address and prefix length are correct.\n")
                    elif GUA in GUAAddress and GUACount == 1 and PrefixLenValue != 64:
                        savereport.write(f"{hostname} IPv6 address is correct but prefix length is not correct.\n")
                    elif GUA in GUAAddress and GUACount > 1:
                        savereport.write(f"{hostname} has multiple IPv6 configured.\n")
                    else:
                        savereport.write(f"{hostname} IPv6 address is not correct.\n")

                    if IntStatus == "up/up":
                        savereport.write(f"{hostname} interface Fa0/0 status is UP.\n")
                    else:
                        savereport.write(f"{hostname} interface Fa0/0 status is DOWN.\n")

            else:
                with open("Lab-Report-Task1.txt", "a") as savereport:
                    savereport.write(f"\n{hostname} IPv6 address is not configured.\n")

    except Exception as e:
        with open("Lab-Report-Task1.txt", "a") as savereport:
            savereport.write(f"\n{hostname} Error: {str(e)}\n")

def task1_check_wrapper(args):
    task1_check(*args)


def main():
    pool = multiprocessing.Pool(processes=len(HostnameList))
    pool.map(task1_check_wrapper, zip(HostnameList, IPList))
    pool.close()
    pool.join()

    print("#### Finished execution of the script ####")


if __name__ == '__main__':
    main()

# Close the report file
savereport.close()

# Show the report content
with open("Lab-Report-Task1.txt", "r") as showreport:
    print(showreport.read())
