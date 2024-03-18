#!C:\Users\abdullah.naser\AppData\Local\Programs\Python\Python312\python.exe
# Your Python script starts here

#import getpass
#import telnetlib
from netmiko import ConnectHandler
import time
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import ipaddress
#from ipaddress import IPv6Interface
#from ipaddress import IPv6Address
#from ipaddress import IPv6Network
from ipaddress import IPv4Interface
#from ipaddress import IPv4Address
#from ipaddress import IPv4Network

#USER = input("Enter your username: ")
#PASSWORD = getpass.getpass()
USER = "apnic"
PASSWORD = "training"
TIMEOUT = 30

def device_access(IP, HOST):
    global ssh
    cisco_devices = {
        'device_type': 'cisco_ios',
        'host': IP,
        'username': 'apnic',
        'password': 'training'
    }

    try:
        with ConnectHandler(**cisco_devices) as ssh:
            #print(f"Executing script on {HOST}...")
            command1 = (f"show runn")
            #command2 = "Y"
            #command3 = "write memory"
            #command = ["show ip int brief", "Y"]
            print ("")
            output1 = ssh.send_command_timing(command1, strip_prompt=False, strip_command=False)
            time.sleep(3)
            print (f"{HOST}: Task 1 configuration restored successfully")

    except Exception as e:
        print (f"Couldn't connect to {HOST}, check manually!")

# Open Workbook and Sheet 
wb = openpyxl.load_workbook("Device-Info.xlsx")
sheet = wb['IPv6_Basic']

# Iterating Function
max_row = sheet.max_row
max_column = sheet.max_column
HostnameList = []
IPList = []

for i in range(2, 3):
    hostname = sheet.cell(row=i, column=1).value
    IPv4_FA60_Temp = sheet.cell(row=i,column=3).value
    IPv4_FA60_IP = IPv4Interface(IPv4_FA60_Temp)
    IPv4_FA60_IP = str(IPv4_FA60_IP.ip)
    HostnameList.append(hostname)
    IPList.append(IPv4_FA60_IP)

savereport = open("Lab-Report-Task3.txt", "wb")


for IP, HOST in zip(IPList, HostnameList):
    device_access(IP, HOST)

'''
for HOST, IP in zip(HostnameList, IPList):
    try:
        print (f"Collecting router configuration for {IP}...")
        tn = telnetlib.Telnet(IP, timeout=TIMEOUT)

        tn.read_until(b"Username: ")
        tn.write(USER.encode('ascii') + b"\n")
        if PASSWORD:
            tn.read_until(b"Password: ")
            tn.write(PASSWORD.encode('ascii') + b"\n")

        #tn.write(b"ping 2001:db8:abc::2 repeat 5\n")
        tn.write(b"show runn\n")
        tn.write(b"exit\n")
        ###print (tn.read_all().decode('ascii'))

        readoutput = tn.read_all()
        saveoutput = open(HOST + "-task2-compare.txt", "w")
        saveoutput.write(readoutput.decode("ascii"))
        saveoutput.write("\n")
    
        text = str(readoutput)
        print(text)

        
        IntStatusStart = "["
        IntStatusEnd = "]"

        if IntStatusStart in text and IntStatusEnd in text:
            IntStatusStartIndex = text.index(IntStatusStart)
            IntStatusEndIndex = text.index(IntStatusEnd)
            IntStatus = text[IntStatusStartIndex+1:IntStatusEndIndex]
        else:
            print ("Couldn't read Interface Status of " + HOST)

        GUAStart = "Global unicast address"
        GUAEnd = "Joined group address"

        if GUAStart in text and GUAEnd in text:
            GUAStartIndex = text.index(GUAStart)
            GUAStartIndexLen = (len(GUAStart))
            GUAEndIndex = text.index(GUAEnd)
            GUAAddress = text[GUAStartIndex+GUAStartIndexLen:GUAEndIndex]
            CIDRIndex = GUAAddress.index("/")
            CIDRvalue = GUAAddress[CIDRIndex+1:CIDRIndex+4]
            PrefixLenValue = (int(CIDRvalue))
        else:
            print ("Couldn't read GUA of " + HOST)
    
        GUA = "2001:DB8:ABC::1"
        GUACount = text.count("subnet is")

        savereport.write("\n")

        if GUA in GUAAddress and GUACount == 1 and PrefixLenValue == 64:
            savereport.write(HOST + " IPv6 address and prefix length are correct.\n")

        elif GUA in GUAAddress and GUACount == 1 and PrefixLenValue != 64:
            savereport.write(HOST + " IPv6 address is correct but prefix length is not correct.\n")

        elif GUA in GUAAddress and GUACount > 1:
            savereport.write(HOST + " has multiple IPv6 configured.\n")

        else:
            savereport.write(HOST + " IPv6 address is not correct.\n")
 
        if IntStatus == "up/up":
            savereport.write(HOST + " interface Fa0/0 status is UP.\n")
        else:
            savereport.write(HOST + " interface Fa0/0 status is DOWN.\n")
        
        saveoutput.close()

    except Exception as e:
        savereport.write("\n")
        savereport.write("Couldn't connect to " + HOST + " , check manually!")
        savereport.write("\n")
'''


savereport.close()
#saveoutput.close()
