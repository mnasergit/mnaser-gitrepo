import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill
#import ipaddress
from ipaddress import IPv4Interface
import base64
#import jinja2
from jinja2 import Environment, FileSystemLoader
#import os

# Open Workbook and Sheet 
wb = openpyxl.load_workbook("Device-Info.xlsx")
sheet = wb['IPv6_Basic']

# Iterating Function
max_row = sheet.max_row
max_column = sheet.max_column
HostnameList = []
IPList = []

for i in range(2, max_row + 1):
    hostname = sheet.cell(row=i, column=1).value
    IPv4_FA60_Temp = sheet.cell(row=i,column=3).value
    IPv4_FA60_IP = IPv4Interface(IPv4_FA60_Temp)
    IPv4_FA60_IP = str(IPv4_FA60_IP.ip)    
    HostnameList.append(hostname)
    IPList.append(IPv4_FA60_IP)
    
for HOST, IP in zip(HostnameList, IPList):    
    # Read the contents of the text file
    with open(f"ipv6_basic_out_files/{HOST}-startup-config.txt", "rb") as file:
        print (f"Reading {HOST}-startup-config.txt")
        text_contents = file.read()
        
        # Encode the contents using base64
        encoded_contents = base64.b64encode(text_contents)

        # Write the encoded contents to a new file
        with open(f"ipv6_basic_out_files/{HOST}-startup-config-hashed.txt", 'wb') as encoded_file:
            print (f"Writing hash for {HOST}-startup-config.txt")
            encoded_file.write(encoded_contents)

        encoded_file.close()
    file.close()

text_contents2 =[]

template_dir = ""
template_file = "Lab1-template.j2"
output_directory = ""
env = Environment(loader=FileSystemLoader(template_dir))
template = env.get_template(template_file)

for HOST, IP in zip(HostnameList, IPList):    
    # Read the contents of the text file
    with open(f"ipv6_basic_out_files/{HOST}-startup-config-hashed.txt", "r") as file2:
        print (f"Reading {HOST}-startup-config-hashed.txt")
        text_contents2.append(file2.read())

dict = {}

i = (len(HostnameList))
for j in range(1, i+1):
    key = f"Group{j}_Router1"
    value = text_contents2[j-1]
    dict[key] = value

data = [dict]

'''        
data = [{
        f"Group{1}_Router1" : text_contents2[0],
        f"Group{2}_Router1" : text_contents2[1],
        f"Group{3}_Router1" : text_contents2[2]
        }]

print(data)
print(type(data))
print(len(data))
'''

for p in data:
    f = open("Lab1-final.unl", "w")
    f.seek(0)
    result = template.render(p)
    f.write(result)

file2.close()
f.close()