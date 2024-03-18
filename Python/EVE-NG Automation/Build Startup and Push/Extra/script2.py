import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import ipaddress
from ipaddress import IPv6Address
from ipaddress import IPv6Network
from ipaddress import IPv6Interface
from ipaddress import IPv4Address
from ipaddress import IPv4Network
from ipaddress import IPv4Interface

wb = openpyxl.load_workbook("Node-Info.xlsx")
sheet = wb['Sheet1']

with open("common_config.txt", 'r') as common_config:
    common_cfg = common_config.readlines()

max_row = sheet.max_row
max_column = sheet.max_column

for i in range(2, max_row + 1):
    hostname = sheet.cell(row=i, column=1).value
    with open(f"{hostname}.txt", 'w') as f:
        f.seek(0)

        f.write("!!! Node Specific Configuration" + "\n")
        f.write("!" + "\n")
        f.write(f' hostname {hostname}' + "\n")
        f.write("!" + "\n")

        IPv6_FA00_Temp = sheet.cell(row=i, column=2).value
        if IPv6_FA00_Temp is not None:
            IPv6_FA00_IP = IPv6Interface(IPv6_FA00_Temp)
            f.write("interface fastEthernet 0/0" + "\n")
            f.write(" description **Connected to Host1**" + "\n")
            f.write(" ipv6 address " + str(IPv6_FA00_IP.ip) + "/" + str(IPv6_FA00_IP._prefixlen) + "\n")
            f.write(" ip address 10.10.10.1 255.255.255.252 !Required to keep interface UP during startup" + "\n")
            f.write(" no shutdown" + "\n")
            f.write("!" + "\n")

        IPv6_FA60_Temp = sheet.cell(row=i, column=3).value
        if IPv6_FA60_Temp is not None:
            IPv6_FA60_IP = IPv6Interface(IPv6_FA60_Temp)
            f.write("interface fastethernet 6/0" + "\n")
            f.write(" description **MGMT - Do Not Delete**" + "\n")
            f.write(" ipv6 address " + str(IPv6_FA60_IP.ip) + "/" + str(IPv6_FA60_IP._prefixlen) + "\n")

        IPv4_FA60_Temp = sheet.cell(row=i, column=4).value
        if IPv4_FA60_Temp is not None:
            IPv4_FA60_IP = IPv4Interface(IPv4_FA60_Temp)
            f.write(" ip address " + str(IPv4_FA60_IP.ip) + " " + str(IPv4_FA60_IP.netmask) + "\n")
            f.write(" no shutdown" + "\n")

        f.write("!" + "\n")
        f.writelines(common_cfg)
        f.write("!" + "\n")
        f.write("end" + "\n")
        f.close()
