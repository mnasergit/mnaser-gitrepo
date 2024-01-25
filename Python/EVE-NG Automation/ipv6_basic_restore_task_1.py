#!C:\Users\abdullah.naser\AppData\Local\Programs\Python\Python312\python.exe
# Your Python script starts here

#import paramiko
from netmiko import ConnectHandler
import time
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import ipaddress
#from ipaddress import IPv6Interface
#from ipaddress import IPv6Address
#from ipaddress import IPv6Network
from ipaddress import IPv4Interface
#from ipaddress import IPv4Address
#from ipaddress import IPv4Network
import requests
import json
from pprint import pprint
import time
import sys
import app
import re
from datetime import datetime


#USER = input("Enter your username: ")
#PASSWORD = getpass.getpass()

EVE_1 = "192.168.20.12"
USER = "apnic"
PASSWORD = "training"
TIMEOUT = 3

### User Input ###

#lab_name = str(input("Enter lab name : "))
lab_name = "IPv6-Basic-Connectivity-Lab"
lab_name_check = f"/{lab_name}.unl"
#num_group = int(input("Enter number Group for Basic IPv6 Connectivity Lab: "))


### Login ###

login_url = f"http://{EVE_1}/api/auth/login"
cred = '{"username":"admin","password":"eve","html5":"-1"}'
headers = {"Accept":"application/json"}

login_api = requests.post(url=login_url, data=cred)
cookies = login_api.cookies

# Check Status of Node #

node_status_url = f"http://{EVE_1}/api/labs/IPv6-Basic-Connectivity-Lab.unl/nodes"
node_status_api = requests.get(url=node_status_url,cookies=cookies,headers=headers)

node_status_api_response = node_status_api.json()
#pprint (node_status_api_response)

node_status_dict = node_status_api_response['data']
num_nodes = len(node_status_dict)


def device_access(IP, HOST):
    global ssh
    cisco_devices = {
        'device_type': 'cisco_ios',
        'host': IP,
        'username': 'apnic',
        'password': 'training'
    }

    try:
        with ConnectHandler(**cisco_devices) as ssh:
            #print(f"Executing script on {HOST}...")
            command1 = (f"configure replace tftp://10.99.99.11/lab1/{HOST}-task1-config.txt")
            command2 = "Y"
            command3 = "write memory"
            #command = ["show ip int brief", "Y"]
            print ("")
            output1 = ssh.send_command_timing(command1, strip_prompt=False, strip_command=False)
            time.sleep(3)
            #print (output1)
            output2 = ssh.send_command_timing(command2, strip_prompt=False, strip_command=False)
            time.sleep(10)
            #print (output2)
            output3 = ssh.send_command_timing(command3, strip_prompt=False, strip_command=False)
            time.sleep(2)
            #print (output3)
            print (f"{HOST}: Task 1 configuration restored successfully")

    except Exception as e:
        print (f"Couldn't connect to {HOST}, check manually!")

# Open Workbook and Sheet
wb = openpyxl.load_workbook("Device-Info.xlsx")
sheet = wb['IPv6_Basic']

# Iterating Function
max_row = sheet.max_row
max_column = sheet.max_column
HostnameList = []
IPList = []

for i in range(2, int(num_nodes/2) + 2):
    hostname = sheet.cell(row=i, column=1).value
    IPv4_FA60_Temp = sheet.cell(row=i,column=3).value
    IPv4_FA60_IP = IPv4Interface(IPv4_FA60_Temp)
    IPv4_FA60_IP = str(IPv4_FA60_IP.ip)
    HostnameList.append(hostname)
    IPList.append(IPv4_FA60_IP)

#command = ["configure replace tftp://10.99.99.11/Group1-Router1-startup-config", "Y"]
#ssh.send_config_set(command)
#ips = ["172.16.1.2", "172.16.1.3"]

for IP, HOST in zip(IPList, HostnameList):
    device_access(IP, HOST)
    #print(f"Executing script on {HOST} {IP}...")
    #restore_config(IP)
    #ssh.send_config_set(command)
    #time.sleep(10)
