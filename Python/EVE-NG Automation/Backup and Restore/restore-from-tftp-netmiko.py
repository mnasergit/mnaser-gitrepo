#import paramiko
from netmiko import ConnectHandler
import time
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import ipaddress
#from ipaddress import IPv6Interface
#from ipaddress import IPv6Address
#from ipaddress import IPv6Network
from ipaddress import IPv4Interface
#from ipaddress import IPv4Address
#from ipaddress import IPv4Network

def device_access(IP, HOST):
    global ssh
    cisco_devices = {
        'device_type': 'cisco_ios',
        'host': IP,
        'username': 'apnic',
        'password': 'training'
    }

    try:
        with ConnectHandler(**cisco_devices) as ssh:
            print(f"Executing script on {HOST}...")
            command1 = (f"configure replace tftp://10.99.99.11/{HOST}-startup-config.txt")
            command2 = "Y"
            command3 = "write memory"
            #command = ["show ip int brief", "Y"]
            output1 = ssh.send_command_timing(command1, strip_prompt=False, strip_command=False)
            time.sleep(2)
            print (output1)
            time.sleep(2)
            output2 = ssh.send_command_timing(command2, strip_prompt=False, strip_command=False)
            time.sleep(2)
            print (output2)
            time.sleep(10)
            output3 = ssh.send_command_timing(command3, strip_prompt=False, strip_command=False)
            time.sleep(2)
            print (output3)
            time.sleep(2)
            #ssh.send_command = (f"configure replace tftp://10.99.99.11/{HOST}-startup-config.txt")
            #ssh.send_command = ("show ip int brief")
            #ssh.send_command = ("Y")
            #time.sleep(15)

    except Exception as e:
        print (f"Couldn't connect to {HOST}, check manually!")

# Open Workbook and Sheet
wb = openpyxl.load_workbook("Device-Info.xlsx")
sheet = wb['IPv6_Basic']

# Iterating Function
max_row = sheet.max_row
max_column = sheet.max_column
HostnameList = []
IPList = []

for i in range(2, max_row + 1):
    hostname = sheet.cell(row=i, column=1).value
    IPv4_FA60_Temp = sheet.cell(row=i,column=3).value
    IPv4_FA60_IP = IPv4Interface(IPv4_FA60_Temp)
    IPv4_FA60_IP = str(IPv4_FA60_IP.ip)
    HostnameList.append(hostname)
    IPList.append(IPv4_FA60_IP)

#command = ["configure replace tftp://10.99.99.11/Group1-Router1-startup-config", "Y"]
#ssh.send_config_set(command)
#ips = ["172.16.1.2", "172.16.1.3"]

for IP, HOST in zip(IPList, HostnameList):
    device_access(IP, HOST)
    #print(f"Executing script on {HOST} {IP}...")
#    restore_config(IP)
    #ssh.send_config_set(command)
    #time.sleep(10)
