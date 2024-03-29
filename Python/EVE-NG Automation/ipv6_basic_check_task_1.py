#!C:\Users\abdullah.naser\AppData\Local\Programs\Python\Python312\python.exe
# Your Python script starts here

#import getpass
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import ipaddress
#from ipaddress import IPv6Interface
#from ipaddress import IPv6Address
#from ipaddress import IPv6Network
from ipaddress import IPv4Interface
#from ipaddress import IPv4Address
#from ipaddress import IPv4Network
import requests
import json
from pprint import pprint
import time
import sys
import app
import re
from datetime import datetime
import threading
from netmiko import ConnectHandler



#USER = input("Enter your username: ")
#PASSWORD = getpass.getpass()
EVE_1 = "192.168.20.12"
EVE_USER = "apnic"
EVE_PASSWORD = "training"
ROUTER_USER = "apnic"
ROUTER_PASSWORD = "training"
TIMEOUT = 6

### User Input ###

#lab_name = str(input("Enter lab name : "))
lab_name = "IPv6-Basic-Connectivity-Lab"
lab_name_check = f"/{lab_name}.unl"
#num_group = int(input("Enter number Group for Basic IPv6 Connectivity Lab: "))


### Login ###

login_url = f"http://{EVE_1}/api/auth/login"
cred = '{"username":"admin","password":"eve","html5":"-1"}'
headers = {"Accept":"application/json"}

login_api = requests.post(url=login_url, data=cred)
cookies = login_api.cookies

# Check Status of Node #

node_status_url = f"http://{EVE_1}/api/labs/IPv6-Basic-Connectivity-Lab.unl/nodes"
node_status_api = requests.get(url=node_status_url,cookies=cookies,headers=headers)

node_status_api_response = node_status_api.json()
#pprint (node_status_api_response)

node_status_dict = node_status_api_response['data']
num_nodes = len(node_status_dict)

# Open Workbook and Sheet 
wb = openpyxl.load_workbook("Device-Info.xlsx")
sheet = wb['IPv6_Basic']

# Iterating Function
max_row = sheet.max_row
max_column = sheet.max_column
HostnameList = []
IPList = []

for i in range(2, int(num_nodes/2)+2):
    hostname = sheet.cell(row=i, column=1).value
    IPv4_FA60_Temp = sheet.cell(row=i,column=3).value
    IPv4_FA60_IP = IPv4Interface(IPv4_FA60_Temp)
    IPv4_FA60_IP = str(IPv4_FA60_IP.ip)    
    HostnameList.append(hostname)
    IPList.append(IPv4_FA60_IP)

savereport = open("Lab-Report-Task1.txt", "w")

for HOST, IP in zip(HostnameList, IPList):
    try:
        global ssh
        cisco_devices = {
            'device_type': 'cisco_ios',
            'host': IP,
            'username': ROUTER_USER,
            'password': ROUTER_PASSWORD
        }

        with ConnectHandler(**cisco_devices) as ssh:
            command1 = (f"show ipv6 interface brief fastEthernet0/0")
            output1 = ssh.send_command_timing(command1, strip_prompt=False, strip_command=False)
            time.sleep(1)
            command2 = (f"show ipv6 interface fastEthernet0/0")
            output2 = ssh.send_command_timing(command2, strip_prompt=False, strip_command=False)
            time.sleep(1)

            # Open the file in write mode and write the output to it
            with open(HOST + "-task1-compare.txt", "w") as saveoutput:
                #saveoutput.write(output1.decode("ascii"))
                saveoutput.write(output1)
                saveoutput.write("\n")
                saveoutput.write(output2)
                saveoutput.write("\n")
                #saveoutput.close()

            with open(HOST + "-task1-compare.txt", "r") as readoutput:
                text = readoutput.read()

            IntStatusStart = "["
            IntStatusEnd = "]"

            if IntStatusStart in text and IntStatusEnd in text:
                IntStatusStartIndex = text.index(IntStatusStart)
                IntStatusEndIndex = text.index(IntStatusEnd)
                IntStatus = text[IntStatusStartIndex+1:IntStatusEndIndex]
            
            else:
                print ("Couldn't read Interface Status of " + HOST)

            GUAStart = "Global unicast address"
            GUAEnd = "Joined group address"

            if GUAStart in text and GUAEnd in text:
                GUAStartIndex = text.index(GUAStart)
                GUAStartIndexLen = (len(GUAStart))
                GUAEndIndex = text.index(GUAEnd)
                GUAAddress = text[GUAStartIndex+GUAStartIndexLen:GUAEndIndex]
                CIDRIndex = GUAAddress.index("/")
                CIDRvalue = GUAAddress[CIDRIndex+1:CIDRIndex+4]
                PrefixLenValue = (int(CIDRvalue))

                GUA = "2001:DB8:ABC::1"
                GUACount = text.count("subnet is")

                savereport.write("\n")

                if GUA in GUAAddress and GUACount == 1 and PrefixLenValue == 64:
                    savereport.write(HOST + " IPv6 address and prefix length are correct.\n")

                elif GUA in GUAAddress and GUACount == 1 and PrefixLenValue != 64:
                    savereport.write(HOST + " IPv6 address is correct but prefix length is not correct.\n")

                elif GUA in GUAAddress and GUACount > 1:
                    savereport.write(HOST + " has multiple IPv6 configured.\n")

                else:
                    savereport.write(HOST + " IPv6 address is not correct.\n")

                if IntStatus == "up/up":
                    savereport.write(HOST + " interface Fa0/0 status is UP.\n")
                else:
                    savereport.write(HOST + " interface Fa0/0 status is DOWN.\n")

            else:
                savereport.write("\n")
                savereport.write(HOST + " IPv6 address is not configured.")
                savereport.write("\n")

    except:
        savereport.write("\n")
        savereport.write(HOST + " Couldn't connect, check manually.")        
        savereport.write("\n")

savereport.close()

showreport = open("Lab-Report-Task1.txt", "r")
showreport = showreport.read()
    
print (showreport)
    
#savereport.close()
#saveoutput.close()
